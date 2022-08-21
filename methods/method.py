import pandas as pd
import numpy as np
from datetime import datetime
from datetime import date
import calendar

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from tempfile import NamedTemporaryFile
from methods.sharepoint import loadShareFile, saveSharedFile

def obtenerDataAgregada(data,sede,pathToSave):
    # columnasPermitidas=['DOC_CORRELATIVO','CENTRO_COSTO','FECHA_CONTABILIZACION','ESTADO_OFERTA','COD_CLIENTE','NOMBRE_CLIENTE','TELEFONO1','EMAIL','VENDEDOR','COD_ARTICULO']
    # cuantos por tipo de asesor
    data=data.rename(columns=data.iloc[0]).drop(data.index[0])  # first row as headers
    data['Cantidad']=1
    # print(data.head(10))
    table = pd.pivot_table(data,values=['Cantidad'], index=['VENDEDOR'], columns=['Estado'], aggfunc=np.sum)
    table.to_excel(pathToSave + sede + ' resumen.xlsx')
    # canto tiempo está en cada estado
    # cuanto es el ratio de conversión 
    pass

def actualizarArchivo(data, sede, pathToSave):
    # label file
    if sede=='Honda SM Retail':
        # wb = load_workbook('prueba proyección SM.xlsx')
        wb = load_workbook(BytesIO(loadShareFile(filename='proyección SM.xlsx')))
    elif sede =='Honda SQ Retail':
        # wb = load_workbook('prueba proyección SQ.xlsx')
        wb = load_workbook(BytesIO(loadShareFile(filename='proyección SQ.xlsx')))
        
    # ws2= wb['base']  
    ws2= wb['proyeccion']
    
    # defining validation data
    dv = DataValidation(type="list", formula1='"Frio,Tibio,Caliente,Perdido,Cerrado"', allow_blank=True)
    dv.error ='No es una opcion válida'
    dv.errorTitle = 'Ventas Motos: Entrada errada'
    dv.prompt = 'Seleccione un estado'
    dv.promptTitle = 'Estado de la Oferta'

    # save previous comments
    dic_comments={}
    n=2
    while ws2.cell(row=n, column=5).value=='Abierto':
        if ws2.cell(row=n, column=12).value!=None:
            dic_comments[str(ws2.cell(row=n, column=2).value)]=[ws2.cell(row=n, column=12).value,ws2.cell(row=n, column=13).value]
        n+=1

    # clear cells 
    n=2
    while ws2.cell(row=n, column=1).value!=None:
        ws2.cell(row=n, column=1).value=None
        ws2.cell(row=n, column=2).value=None
        ws2.cell(row=n, column=3).value=None
        ws2.cell(row=n, column=4).value=None
        ws2.cell(row=n, column=5).value=None
        ws2.cell(row=n, column=6).value=None
        ws2.cell(row=n, column=7).value=None
        ws2.cell(row=n, column=8).value=None
        ws2.cell(row=n, column=9).value=None
        ws2.cell(row=n, column=10).value=None
        ws2.cell(row=n, column=11).value=None
        ws2.cell(row=n, column=12).value=None
        ws2.cell(row=n, column=13).value=None
        n+=1

    # update
    n=2
    for row in data.itertuples():
        if row.CENTRO_COSTO==sede:
            ws2.cell(row=n, column=1, value=row.Index)
            ws2.cell(row=n, column=2, value=row.DOC_CORRELATIVO)
            ws2.cell(row=n, column=3, value=row.CENTRO_COSTO)
            ws2.cell(row=n, column=4, value=row.FECHA_CONTABILIZACION)
            ws2.cell(row=n, column=5, value=row.ESTADO_OFERTA)
            ws2.cell(row=n, column=6, value=row.COD_CLIENTE)
            ws2.cell(row=n, column=7, value=row.NOMBRE_CLIENTE)
            ws2.cell(row=n, column=8, value=row.TELEFONO1)
            ws2.cell(row=n, column=9, value=row.EMAIL)
            ws2.cell(row=n, column=10, value=row.VENDEDOR)
            ws2.cell(row=n, column=11, value=row.COD_ARTICULO)
            n=n+1

    #  update comments
    n=2
    print(dic_comments)
    while ws2.cell(row=n, column=5).value=='Abierto':
        try:
            ws2.cell(row=n, column=12, value=dic_comments[str(ws2.cell(row=n, column=2).value)][0])
            ws2.cell(row=n, column=13, value=dic_comments[str(ws2.cell(row=n, column=2).value)][1])
        except:
            pass
        n+=1

    # set validation data
    ws2.add_data_validation(dv)
    dv.add(f'L2:L{data.shape[0]}')  # colocar validacion hasta la ultima necesaria
    
    # openpyxl -> dataframe pivot table
    obtenerDataAgregada(pd.DataFrame(ws2.values), sede, pathToSave)

    # save
    if sede=='Honda SM Retail':
        wb.save(filename = pathToSave +'proyección SM.xlsx')
    elif sede =='Honda SQ Retail':
        wb.save(filename = pathToSave +'proyección SQ.xlsx')

    print(f'{sede} updated')

def prepararData(file_name):
    df=pd.read_excel(file_name)
    
    # solo ofertas abiertas (revisar esto luego)
    df= df[df['ESTADO_OFERTA']=='Abierto']

    # solo unidades motos
    df= df[df['GRUPO_ARTICULO']=='UNIDADES MOTOS HONDA']
   
    # solo retail
    valoresPermitidos=['Honda SM Retail','Honda SQ Retail']
    df= df[df['CENTRO_COSTO'].isin(valoresPermitidos)]

    # columnas permitidas
    columnasPermitidas=['DOC_CORRELATIVO','CENTRO_COSTO','FECHA_CONTABILIZACION','ESTADO_OFERTA','COD_CLIENTE','NOMBRE_CLIENTE','TELEFONO1','EMAIL','VENDEDOR','COD_ARTICULO']
    df=df[columnasPermitidas]
    df['FECHA_CONTABILIZACION'] = df['FECHA_CONTABILIZACION'].dt.strftime('%Y%m%d')

    # # solo mes actual
    currentMonth = datetime.now().month  # int
    currentYear = datetime.now().year  # int
    _,num_days = calendar.monthrange(2016, 3)  # tuple
    first_day = date(currentYear, currentMonth, 1)  # date
    last_day = date(currentYear, currentMonth, num_days)  # date
    df['FECHA_CONTABILIZACION'] = pd.to_datetime(df['FECHA_CONTABILIZACION'], format='%Y%m%d')  # datetime
    df=df[ (df['FECHA_CONTABILIZACION']>=pd.to_datetime(first_day)) & (df['FECHA_CONTABILIZACION']<=pd.to_datetime(last_day))]
    df['FECHA_CONTABILIZACION'] = df['FECHA_CONTABILIZACION'].dt.strftime('%Y-%m-%d')  # string
        
    # de mas reciente a mas antiguo
    df.sort_values(by=['DOC_CORRELATIVO'], ascending=False, inplace=True, ignore_index=True)
    return df